from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def apply_formatting(worksheet):
    # Font definitions
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    content_font = Font(name='Calibri', size=11, color="000000")

    # Fill definitions
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Apply header formatting
    for col in range(1, worksheet.max_column + 1):
        header_cell = worksheet.cell(row=1, column=col)
        if header_cell.column_letter == 'A':  # If it's the date column header
            header_cell.font = Font(name='Calibri', size=11, bold=True, color="000000")  # Bold black font for the "Date" header
        else:
            header_cell.font = header_font

        # Apply header fill based on column
        if "B" <= get_column_letter(col) <= "E":
            header_cell.fill = blue_fill
        elif "F" <= get_column_letter(col) <= "I":
            header_cell.fill = orange_fill
        

    # Apply content formatting
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = content_font

    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Adjust date column (if required)
    worksheet.column_dimensions['A'].width = 10 
    worksheet.sheet_view.showGridLines = True